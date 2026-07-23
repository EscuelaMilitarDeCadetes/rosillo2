"""
Réplica de MainController.exportToExcel / exportToPdf del Thymeleaf original
(INV-07). Exporta el mismo resultado que ya arma
ProyectoXConvocatoriaSelector.buscar_con_filtros() (Hallazgo C), en vez de
depender de variables de sesión como hacía el original.

El % de avance se calcula con AvanceService.calcular_avance_ponderado(),
la versión YA CORREGIDA (ver AvanceService), no con la fórmula del Thymeleaf
que, según quedó documentado ahí, no ponderaba de verdad pese al nombre.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font as XlFont, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from apps.institucional.models import PersonaXGrupo
from apps.investigacion_formal.selectors.investigador_x_proyecto_selector import (
    InvestigadorXProyectoSelector,
)
from apps.investigacion_formal.selectors.producto_x_proyecto_selector import (
    ProductoXProyectoSelector,
)
from apps.investigacion_formal.selectors.monto_selector import MontoSelector
from apps.investigacion_formal.services.avance_service import AvanceService

ENCABEZADOS = [
    "Código", "Facultad/Grupo", "Proyecto", "Inicio", "Fin", "% Avance",
    "Convocatoria", "Valor asignado", "Valor contrapartida", "Valor Total",
    "Valor asignado ejecutado", "% Ejecutado", "grupLAC",
    "Investigadores", "Productos",
]


class ExportacionService:

    # ------------------------------------------------------------------
    # Armado de filas (compartido entre Excel y PDF)
    # ------------------------------------------------------------------
    @staticmethod
    def _facultad_grupo_label(usuario_id):
        pxg = (
            PersonaXGrupo.objects
            .filter(
                persona__asignaciones__usuario_id=usuario_id,
                persona__asignaciones__estado=True,
                estado=True,
            )
            .select_related('facultad', 'grupo')
            .first()
        )
        if pxg is None:
            return "N/A"
        if pxg.facultad_id:
            return pxg.facultad.abreviatura
        if pxg.grupo_id:
            return pxg.grupo.sigla_grupo
        return "N/A"

    @staticmethod
    def _formatear_moneda(valor):
        if valor is None or valor <= 0:
            return "N/A"
        return f"${valor:,.2f}"

    @staticmethod
    def _formatear_fecha(fecha):
        return fecha.strftime("%d-%m-%Y") if fecha else "N/A"

    @staticmethod
    def _armar_fila(registro):
        proyecto = registro.proyecto
        monto = MontoSelector.obtener_por_proyecto(proyecto.pk)

        aprobado = monto.aprobado if monto else 0
        ejecutado = monto.ejecutado if monto else 0
        porcentaje_ejecutado = max((ejecutado * 100 / aprobado), 0) if aprobado else 0

        investigadores = InvestigadorXProyectoSelector.listar_por_proyecto(
            proyecto.pk, solo_activos=True
        )
        productos = ProductoXProyectoSelector.listar_por_proyecto(
            proyecto.pk, solo_activos=True
        )

        texto_investigadores = "; ".join(
            f"* {i.rol_investigador.nombre_rol_investigador} - "
            f"{i.persona_x_grupo.persona.nombre} {i.persona_x_grupo.persona.apellido}"
            for i in investigadores
        ) or "Sin investigadores"

        texto_productos = "; ".join(
            f"* {p.producto_x_grupo.grupo_minciencias.nombre_grupo_minciencias} - "
            f"{p.producto_x_grupo.producto_minciencias.nomenclatura} - "
            f"Entregado: {'Sí' if p.entregado else 'No'}"
            for p in productos
        ) or "Sin productos"

        avance_ponderado = AvanceService.calcular_avance_ponderado(proyecto.pk)

        return [
            proyecto.codigo or "Sin código",
            ExportacionService._facultad_grupo_label(proyecto.usuario_id),
            proyecto.titulo,
            ExportacionService._formatear_fecha(proyecto.fecha_inicio),
            ExportacionService._formatear_fecha(proyecto.fecha_fin),
            f"{avance_ponderado} %",
            "Interno" if registro.convocatoria.interno else "Externo",
            ExportacionService._formatear_moneda(monto.aprobado if monto else None),
            ExportacionService._formatear_moneda(monto.contrapartida if monto else None),
            ExportacionService._formatear_moneda(monto.total if monto else None),
            ExportacionService._formatear_moneda(monto.ejecutado if monto else None),
            f"{porcentaje_ejecutado:.2f}%",
            "SI" if proyecto.gruplac else "NO",
            texto_investigadores,
            texto_productos,
        ]

    # ------------------------------------------------------------------
    # Excel
    # ------------------------------------------------------------------
    @staticmethod
    def exportar_excel(queryset):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Proyectos"

        sheet.append(ENCABEZADOS)
        for celda in sheet[1]:
            celda.font = XlFont(bold=True, color="FFFFFF")
            celda.fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")

        for registro in queryset:
            sheet.append(ExportacionService._armar_fila(registro))

        for columna in sheet.columns:
            longitud = max(len(str(c.value)) if c.value is not None else 0 for c in columna)
            sheet.column_dimensions[columna[0].column_letter].width = min(longitud + 2, 60)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # ------------------------------------------------------------------
    # PDF
    # ------------------------------------------------------------------
    @staticmethod
    def exportar_pdf(queryset):
        buffer = BytesIO()
        documento = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=10 * mm, rightMargin=10 * mm,
            topMargin=10 * mm, bottomMargin=10 * mm,
        )
        estilos = getSampleStyleSheet()
        titulo_estilo = ParagraphStyle(
            "TituloExport", parent=estilos["Heading1"], alignment=1, fontSize=16,
        )

        celda_estilo = ParagraphStyle("CeldaExport", parent=estilos["Normal"], fontSize=6.5, leading=8)
        encabezado_estilo = ParagraphStyle(
            "EncabezadoExport", parent=estilos["Normal"], fontSize=7.5, leading=9,
            textColor=colors.white,
        )

        datos = [[Paragraph(h, encabezado_estilo) for h in ENCABEZADOS]]
        for registro in queryset:
            fila = ExportacionService._armar_fila(registro)
            datos.append([Paragraph(str(valor), celda_estilo) for valor in fila])

        anchos = [22, 18, 40, 18, 18, 15, 15, 22, 22, 22, 22, 15, 12, 45, 45]
        escala = (landscape(A4)[0] - 20 * mm) / sum(anchos)
        anchos_mm = [a * escala for a in anchos]

        tabla = Table(datos, colWidths=anchos_mm, repeatRows=1)
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#404040")),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ]))

        documento.build([
            Paragraph("Listado de Proyectos", titulo_estilo),
            Spacer(1, 6 * mm),
            tabla,
        ])
        buffer.seek(0)
        return buffer