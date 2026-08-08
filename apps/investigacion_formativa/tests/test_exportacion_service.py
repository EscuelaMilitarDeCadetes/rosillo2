#apps/investigacion_formativa/tests/test_exportacion_service.py
from django.test import TestCase
from openpyxl import load_workbook
from io import BytesIO

from .base import InvestigacionFormativaFixturesMixin
from apps.investigacion_formativa.services.exportacion_service import ExportacionService
from apps.investigacion_formativa.selectors.proceso_formativo_selector import (
    ProcesoFormativoSelector,
)


class ExportacionServiceTests(InvestigacionFormativaFixturesMixin, TestCase):
    """
    self.proceso (del fixture) ya tiene un participante ESTUDIANTE
    (self.participante, con self.persona) y pertenece a self.modalidad,
    disponible en self.facultad vía self.modalidad_facultad — suficiente
    para ejercitar _armar_fila() sin más setup.
    """

    def test_exportar_excel_genera_libro_valido_con_una_fila_por_proceso(self):
        queryset = ProcesoFormativoSelector.listar().filter(pk=self.proceso.pk)

        buffer = ExportacionService.exportar_excel(queryset)

        self.assertIsInstance(buffer, BytesIO)
        wb = load_workbook(buffer)
        sheet = wb.active
        # fila 1 = encabezados, fila 2 = self.proceso
        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet.cell(row=1, column=1).value, "ID")
        self.assertEqual(sheet.cell(row=2, column=2).value, self.proceso.titulo)

    def test_exportar_pdf_genera_documento_valido(self):
        queryset = ProcesoFormativoSelector.listar().filter(pk=self.proceso.pk)

        buffer = ExportacionService.exportar_pdf(queryset)

        self.assertIsInstance(buffer, BytesIO)
        contenido = buffer.read()
        self.assertTrue(contenido.startswith(b"%PDF"))

    def test_exportar_excel_sin_resultados_solo_trae_encabezados(self):
        queryset = ProcesoFormativoSelector.listar().none()

        buffer = ExportacionService.exportar_excel(queryset)

        wb = load_workbook(buffer)
        self.assertEqual(wb.active.max_row, 1)

    def test_armar_fila_sin_facultad_asociada_retorna_na(self):
        # self.persona (la estudiante del fixture) no tiene PersonaXGrupo
        # creado en el fixture base, así que _facultad_label debe caer al
        # "N/A" en vez de reventar.
        fila = ExportacionService._armar_fila(self.proceso)
        indice_facultad = 3  # ["ID","Título","Modalidad","Facultad", ...]
        self.assertEqual(fila[indice_facultad], "N/A")


class BuscarConFiltrosTests(InvestigacionFormativaFixturesMixin, TestCase):

    def test_filtra_por_modalidad(self):
        resultados = ProcesoFormativoSelector.buscar_con_filtros(
            modalidad=self.modalidad.pk
        )
        self.assertIn(self.proceso, list(resultados))

    def test_filtra_por_facultad(self):
        resultados = ProcesoFormativoSelector.buscar_con_filtros(
            facultad=self.facultad.pk
        )
        self.assertIn(self.proceso, list(resultados))

    def test_filtra_por_requiere_sustentacion_excluye_no_coincidentes(self):
        otro_proceso = self._crear_proceso_formativo(titulo='Otro proceso')
        otro_proceso.requiere_sustentacion = False
        otro_proceso.save(update_fields=['requiere_sustentacion'])

        self.proceso.requiere_sustentacion = True
        self.proceso.save(update_fields=['requiere_sustentacion'])

        resultados = list(
            ProcesoFormativoSelector.buscar_con_filtros(requiere_sustentacion=True)
        )
        self.assertIn(self.proceso, resultados)
        self.assertNotIn(otro_proceso, resultados)

    def test_filtra_por_persona_participante(self):
        resultados = ProcesoFormativoSelector.buscar_con_filtros(
            persona=self.persona.pk
        )
        self.assertIn(self.proceso, list(resultados))

    def test_sin_filtros_retorna_todos(self):
        resultados = ProcesoFormativoSelector.buscar_con_filtros()
        self.assertIn(self.proceso, list(resultados))

    def test_filtro_sin_coincidencias_retorna_vacio(self):
        resultados = ProcesoFormativoSelector.buscar_con_filtros(
            estado_general='ESTADO_QUE_NO_EXISTE'
        )
        self.assertEqual(list(resultados), [])