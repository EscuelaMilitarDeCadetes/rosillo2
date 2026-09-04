# E:\PROYECTO_ROSILLO\django_react\django\rosillo\apps\investigacion_formativa\services\exportacion_service.py
"""
Equivalente de investigacion_formal/services/exportacion_service.py para
investigacion_formativa. No es una réplica de Thymeleaf (ese módulo no
existía en el sistema legado — ver ProcesoFormativoSelector.buscar_con_filtros),
sino la primera implementación, siguiendo el mismo patrón: exporta el mismo
resultado que arma ProcesoFormativoSelector.buscar_con_filtros(), en vez de
depender de variables de sesión.

El % de avance se lee directamente de ProcesoFormativo.porcentaje_avance
(campo persistido, actualizado por AvanceService.actualizar_porcentaje_avance),
no se recalcula aquí — a diferencia de investigacion_formal, donde Proyecto
no almacena el avance y hay que llamar AvanceService.calcular_avance_ponderado()
por cada fila.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font as XlFont, PatternFill

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from apps.institucional.models import PersonaXGrupo
from apps.investigacion_formativa.selectors.participante_proceso_selector import (
    ParticipanteProcesoSelector,
)

ENCABEZADOS = [
    "ID", "Título", "Modalidad", "Facultad", "Inicio", "Fin", "% Avance",
    "Estado", "Aprobado", "Horas acumuladas", "Requiere sustentación",
    "Participantes",
]


class ExportacionService:

    # ------------------------------------------------------------------
    # Armado de filas (compartido entre Excel y PDF)
    # ------------------------------------------------------------------

    @staticmethod
    def _facultad_label(persona_id):
        """
        A diferencia de investigacion_formal (donde Proyecto.usuario_id
        identifica directamente a quien ejecuta el proyecto),
        ProcesoFormativo no tiene un 'dueño' único: la facultad se toma de
        la persona con rol ESTUDIANTE en 'participantes' (ver
        _persona_estudiante), consultando PersonaXGrupo directamente por
        persona_id (sin pasar por Usuario/asignaciones, porque
        ParticipanteProceso ya guarda la Persona, no el Usuario).
        """
        if persona_id is None:
            return "N/A"
        pxg = (
            PersonaXGrupo.objects
            .filter(persona_id=persona_id, estado=True)
            .select_related('facultad', 'grupo')
            .first()
        )
        if pxg is None:
            return "N/A"
        if pxg.facultad_id:
            return pxg.facultad.abreviatura
        if pxg.grupo_id:
            return pxg.grupo.sigla_grupo
        return "N/A"

    @staticmethod
    def _persona_estudiante(proceso_id):
        """Persona con rol ESTUDIANTE activo en el proceso (puede no haber
        ninguno todavía si el proceso está en una etapa temprana)."""
        participante = (
            ParticipanteProcesoSelector.listar_estudiantes_por_proceso(proceso_id)
            .filter(activo=True)
            .select_related('persona')
            .first()
        )
        return participante.persona if participante else None

    @staticmethod
    def _formatear_fecha(fecha):
        return fecha.strftime("%d-%m-%Y") if fecha else "N/A"

    @staticmethod
    def _armar_fila(proceso):
        estudiante = ExportacionService._persona_estudiante(proceso.pk)
        facultad_label = ExportacionService._facultad_label(
            estudiante.pk if estudiante else None
        )
        participantes = ParticipanteProcesoSelector.listar_activos_por_proceso(proceso.pk)
        texto_participantes = "; ".join(
            f"* {p.get_rol_en_modalidad_display()} - {p.persona.nombre} {p.persona.apellido}"
            for p in participantes
        ) or "Sin participantes"

        return [
            proceso.pk,
            proceso.titulo or "Sin título",
            proceso.flujo_version.modalidad.nombre if proceso.flujo_version_id else "N/A",
            facultad_label,
            ExportacionService._formatear_fecha(proceso.fecha_inicio),
            ExportacionService._formatear_fecha(proceso.fecha_fin),
            f"{proceso.porcentaje_avance or 0} %",
            proceso.estado_actual,
            "SI" if proceso.aprobado else ("NO" if proceso.aprobado is False else "Pendiente"),
            f"{proceso.horas_acumuladas or 0}",
            "SI" if proceso.requiere_sustentacion else "NO",
            texto_participantes,
        ]

    # ------------------------------------------------------------------
    # Excel
    # ------------------------------------------------------------------

    @staticmethod
    def exportar_excel(queryset):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Procesos Formativos"
        sheet.append(ENCABEZADOS)
        for celda in sheet[1]:
            celda.font = XlFont(bold=True, color="FFFFFF")
            celda.fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
        for proceso in queryset:
            sheet.append(ExportacionService._armar_fila(proceso))
        for columna in sheet.columns:
            longitud = max(len(str(c.value)) if c.value is not None else 0 for c in columna)
            sheet.column_dimensions[columna[0].column_letter].width = min(longitud + 2, 60)
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # ------------------------------------------------------------------
    # PDF
    # ------------------------------------------------------------------

    @staticmethod
    def exportar_pdf(queryset):
        buffer = BytesIO()
        documento = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=10 * mm, rightMargin=10 * mm,
            topMargin=10 * mm, bottomMargin=10 * mm,
        )
        estilos = getSampleStyleSheet()
        titulo_estilo = ParagraphStyle(
            "TituloExport", parent=estilos["Heading1"], alignment=1, fontSize=16,
        )
        celda_estilo = ParagraphStyle("CeldaExport", parent=estilos["Normal"], fontSize=6.5, leading=8)
        encabezado_estilo = ParagraphStyle(
            "EncabezadoExport", parent=estilos["Normal"], fontSize=7.5, leading=9,
            textColor=colors.white,
        )
        datos = [[Paragraph(h, encabezado_estilo) for h in ENCABEZADOS]]
        for proceso in queryset:
            fila = ExportacionService._armar_fila(proceso)
            datos.append([Paragraph(str(valor), celda_estilo) for valor in fila])
        anchos = [10, 40, 25, 18, 18, 18, 15, 18, 15, 18, 18, 50]
        escala = (landscape(A4)[0] - 20 * mm) / sum(anchos)
        anchos_mm = [a * escala for a in anchos]
        tabla = Table(datos, colWidths=anchos_mm, repeatRows=1)
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#404040")),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ]))
        documento.build([
            Paragraph("Listado de Procesos Formativos", titulo_estilo),
            Spacer(1, 6 * mm),
            tabla,
        ])
        buffer.seek(0)
        return buffer